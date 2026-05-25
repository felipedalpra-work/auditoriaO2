"""
O2 Inc — Gerador de Relatório Excel de Auditoria
Aba 1: Resumo Executivo | Abas seguintes: uma por regra de auditoria
"""
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd

# ─── Paleta O2 (sem #) ────────────────────────────────────────────────────────
C_VERDE       = '6CF269'
C_CINZA_ESC   = '494949'
C_CINZA_MED   = '6B6B6B'
C_CINZA_CLARO = 'F4F4F4'
C_BRANCO      = 'FFFFFF'
C_TEXTO       = '222222'

C_CRITICO = 'C0392B'
C_ALTO    = 'E67E22'
C_MEDIO   = 'B8860B'
C_BAIXO   = '27AE60'

SEV_COR = {'CRÍTICO': C_CRITICO, 'ALTO': C_ALTO, 'MÉDIO': C_MEDIO, 'BAIXO': C_BAIXO}
SEV_BG  = {'CRÍTICO': 'FDF0EF', 'ALTO': 'FEF7EF', 'MÉDIO': 'FDFBEF', 'BAIXO': 'EFF8F2'}

# ─── Nomes curtos para abas (máx 31 chars, sem [ ] : * ? / \) ────────────────
_TAB_MAP = {
    'Data de Emissão > Data de Vencimento (+60 dias)': 'Emissão x Venc. 60d+',
    'Data de Vencimento = Data de Emissão':   'Vencimento = Emissão',
    'Data de Vencimento > Data de Emissão (+60 dias)': 'Vencimento > Emissão 60d+',
    'Sinal de Valor Incorreto':               'Sinal Incorreto',
    'Lançamento Duplicado':                   'Duplicados',
    'Recorrência Quebrada':                   'Recorrência Quebrada',
    'Categoria Inconsistente por Fornecedor': 'Categoria Inconsistente',
    'Valor Atípico (Outlier)':                'Valor Atípico',
    'Parcelas Incompletas':                   'Parcelas Incompletas',
    'Competência x Registro Distantes':       'Competência x Registro',
    'Data de Registro Replicada (Bug Omie)':  'Registro Replicado',
    'Juros não Separados da Amortização':     'Juros x Amortização',
    'Categoria Semanticamente Incorreta':     'Categoria Incorreta (IA)',
}

EXPLICACOES = {
    'Data de Emissão > Data de Vencimento (+60 dias)': {
        'o_que_e': (
            'No arquivo origem, a data de emissão está mais de 60 dias após a data de vencimento.'
        ),
        'por_que_importa': (
            'Diferença muito alta entre vencimento e emissão pode indicar erro de competência '
            'ou cadastro incorreto de datas no lançamento.'
        ),
        'como_corrigir': (
            'Consultar o documento fiscal original e corrigir a data de emissão ou '
            'vencimento no Omie antes da próxima importação.'
        ),
    },
    'Data de Vencimento = Data de Emissão': {
        'o_que_e': (
            'No arquivo origem, a data de vencimento é igual à data de emissão. '
            'Com competência pela emissão, isso pode indicar ausência de prazo financeiro.'
        ),
        'por_que_importa': (
            'Concentra reconhecimento e vencimento no mesmo dia e pode mascarar '
            'condições reais de pagamento no fluxo de caixa.'
        ),
        'como_corrigir': (
            'Confirmar no documento/contrato se não há prazo. Se houver, ajustar a data '
            'de vencimento no Omie antes da próxima importação.'
        ),
    },
    'Data de Vencimento > Data de Emissão (+60 dias)': {
        'o_que_e': (
            'No arquivo origem, a data de vencimento está mais de 60 dias após a data de emissão.'
        ),
        'por_que_importa': (
            'Prazo muito longo pode indicar condição fora da política financeira '
            'ou erro de lançamento.'
        ),
        'como_corrigir': (
            'Validar no documento/contrato o prazo de pagamento. '
            'Se incorreto, ajustar a data de vencimento no Omie.'
        ),
    },
    'Sinal de Valor Incorreto': {
        'o_que_e': (
            'Conta a Pagar com valor positivo, ou Conta a Receber com valor negativo. '
            'O sinal indica a direção do fluxo e está invertido.'
        ),
        'por_que_importa': (
            'Uma despesa com sinal positivo aparece como receita na Oxy — distorce '
            'completamente o DRE e análises de receita vs. despesa.'
        ),
        'como_corrigir': (
            'Corrigir o sinal no Omie. Contas a Pagar → valores negativos; '
            'Contas a Receber → valores positivos.'
        ),
    },
    'Lançamento Duplicado': {
        'o_que_e': (
            'Dois ou mais lançamentos com exatamente o mesmo CNPJ/CPF, valor, '
            'data de emissão e número de parcela.'
        ),
        'por_que_importa': (
            'Cada duplicata multiplica o valor real no DRE, distorcendo análises '
            'de custo por categoria, fornecedor e período.'
        ),
        'como_corrigir': (
            'Identificar qual lançamento é correto e excluir os demais no Omie. '
            'Verificar se a importação foi executada mais de uma vez para o mesmo documento.'
        ),
    },
    'Recorrência Quebrada': {
        'o_que_e': (
            'Fornecedor recorrente (presente em >50% dos meses) está ausente '
            'em meses específicos do período analisado.'
        ),
        'por_que_importa': (
            'Despesas fixas (aluguel, assinaturas, folha) devem aparecer em todos os meses. '
            'A ausência pode indicar lançamento no mês errado ou despesa não registrada.'
        ),
        'como_corrigir': (
            'Verificar se o pagamento ocorreu. Se sim, localizar e corrigir o lançamento. '
            'Se não, avaliar e documentar a justificativa.'
        ),
    },
    'Categoria Inconsistente por Fornecedor': {
        'o_que_e': (
            'O mesmo fornecedor foi classificado em categorias diferentes ao longo do período.'
        ),
        'por_que_importa': (
            'A Oxy agrupa despesas por categoria para o DRE. Inconsistências invalidam '
            'comparações entre meses e geram totais incorretos por categoria.'
        ),
        'como_corrigir': (
            'Definir a categoria correta e padronizar todos os lançamentos históricos. '
            'Configurar regra de categorização automática no Omie para este fornecedor.'
        ),
    },
    'Valor Atípico (Outlier)': {
        'o_que_e': (
            'Valor desvia mais de 3 desvios-padrão da média histórica do próprio fornecedor '
            '(z-score > 3). O padrão de gasto habitual com aquele fornecedor é a referência.'
        ),
        'por_que_importa': (
            'Pode indicar erro de digitação, duplicação parcial ou cobrança extraordinária. '
            'Distorce o custo médio do fornecedor e análises de tendência de gastos.'
        ),
        'como_corrigir': (
            'Confirmar o valor no documento fiscal original. Se erro, corrigir no Omie. '
            'Se legítimo (ex: reajuste, serviço extra), documentar na observação do lançamento.'
        ),
    },
    'Parcelas Incompletas': {
        'o_que_e': (
            'Série de parcelas numeradas (ex: 1/12 a 12/12) com lacunas — '
            'parcelas esperadas não foram encontradas no período.'
        ),
        'por_que_importa': (
            'Indica parcelas em meses incorretos ou não lançadas. Compromete '
            'projeções de fluxo de caixa e custo total do contrato no DRE.'
        ),
        'como_corrigir': (
            'Localizar as parcelas faltantes no Omie. Se em datas erradas, '
            'corrigir o período de competência.'
        ),
    },
    'Competência x Registro Distantes': {
        'o_que_e': (
            'Data de emissão e data de registro no Omie distam mais de 60 dias.'
        ),
        'por_que_importa': (
            'A Oxy usa a data de registro como competência. Diferença >60 dias '
            'significa que a despesa está reconhecida no mês errado do DRE.'
        ),
        'como_corrigir': (
            'Corrigir a data de registro no Omie para o mês em que o serviço '
            'foi prestado ou o bem entregue.'
        ),
    },
    'Data de Registro Replicada (Bug Omie)': {
        'o_que_e': (
            'Todas as parcelas da série têm a mesma data de registro — bug do Omie '
            'que replica a data da primeira parcela para todas as demais.'
        ),
        'por_que_importa': (
            'Como a Oxy usa a data de registro como competência, todas as parcelas '
            'ficam no mesmo mês, criando picos artificiais no DRE.'
        ),
        'como_corrigir': (
            'Corrigir manualmente a data de registro de cada parcela no Omie. '
            'Chamado técnico já aberto na Omie para correção do bug.'
        ),
    },
    'Categoria Semanticamente Incorreta': {
        'o_que_e': (
            'O nome do fornecedor e a categoria do lançamento são semanticamente incompatíveis — '
            'o tipo de serviço ou produto indicado pelo fornecedor não corresponde à categoria '
            'em que o lançamento foi registrado.'
        ),
        'por_que_importa': (
            'Lançamentos na categoria errada distorcem o DRE por categoria, '
            'comprometem análises de custo por centro de resultado e podem '
            'gerar inconsistências em relatórios gerenciais e fiscais.'
        ),
        'como_corrigir': (
            'Verificar a natureza do serviço ou produto do fornecedor e corrigir '
            'a categoria no Omie para refletir corretamente o tipo de despesa. '
            'Considerar configurar regra de categorização automática no Omie para este fornecedor.'
        ),
    },
    'Juros não Separados da Amortização': {
        'o_que_e': (
            'Pagamento de empréstimo/financiamento lançado em uma única categoria, '
            'sem separar juros da amortização do principal.'
        ),
        'por_que_importa': (
            'Juros impactam o resultado; amortização não. Lançar tudo como despesa '
            'distorce EBITDA e resultado financeiro.'
        ),
        'como_corrigir': (
            'Separar em dois lançamentos: (1) juros → Despesas Financeiras; '
            '(2) amortização → Empréstimos/Financiamentos. '
            'Valores no extrato do contrato com a instituição financeira.'
        ),
    },
}


# ─── Helpers de estilo ────────────────────────────────────────────────────────

def _fill(c):
    return PatternFill('solid', fgColor=c)

def _font(bold=False, color=C_TEXTO, size=10):
    return Font(bold=bold, color=color, size=size, name='Calibri')

def _align(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border(color='DDDDDD'):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _tab_name(nome_regra):
    if nome_regra in _TAB_MAP:
        return _TAB_MAP[nome_regra]
    sanitized = nome_regra
    for ch in ['[', ']', ':', '*', '?', '/', '\\']:
        sanitized = sanitized.replace(ch, '')
    return sanitized[:31]


def _criar_aba_base_analisada(wb, source_path):
    if not source_path:
        return None
    try:
        base_df = pd.read_excel(source_path, sheet_name=0, header=None)
    except Exception:
        return None

    ws = wb.create_sheet('Base Analisada')
    ws.sheet_view.showGridLines = True

    for r_idx, row_vals in enumerate(base_df.itertuples(index=False, name=None), start=1):
        for c_idx, val in enumerate(row_vals, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    ws.freeze_panes = ws['A2']
    return ws.title


# ─── Aba de Resumo Executivo ──────────────────────────────────────────────────

def _criar_resumo(wb, resultado, nome_arquivo, empresa):
    ws = wb.create_sheet('Resumo Executivo')
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 44
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 16

    total       = resultado['total_lancamentos']
    total_erros = resultado['total_erros']
    pct         = (total_erros / total * 100) if total > 0 else 0
    sev         = resultado.get('por_severidade', {})
    por_regra   = resultado.get('por_regra', {})
    data_analise = resultado.get('data_analise', '')

    row = 1

    # Cabeçalho principal
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.row_dimensions[row].height = 44
    c = ws.cell(row=row, column=1,
                value='O2INC — RELATÓRIO DE AUDITORIA DE LANÇAMENTOS OMIE')
    c.fill = _fill(C_CINZA_ESC); c.font = _font(bold=True, color=C_VERDE, size=14)
    c.alignment = _align('left', 'center')
    row += 1

    # Metadados
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.row_dimensions[row].height = 20
    partes = []
    if empresa:     partes.append(f'Empresa: {empresa}')
    if nome_arquivo: partes.append(f'Arquivo: {nome_arquivo}')
    partes.append(f'Gerado em: {data_analise}')
    c = ws.cell(row=row, column=1, value='  |  '.join(partes))
    c.fill = _fill(C_CINZA_ESC); c.font = _font(color='AAAAAA', size=9)
    c.alignment = _align('left', 'center')
    row += 2  # spacer

    # ── Seção: Indicadores Gerais ─────────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.row_dimensions[row].height = 24
    c = ws.cell(row=row, column=1, value='  INDICADORES GERAIS')
    c.fill = _fill(C_CINZA_ESC); c.font = _font(bold=True, color=C_BRANCO, size=10)
    c.alignment = _align('left', 'center')
    row += 1

    # Header da tabela
    ws.row_dimensions[row].height = 18
    for col, titulo in [(1, 'Indicador'), (2, 'Quantidade'), (3, '% do Total')]:
        c = ws.cell(row=row, column=col, value=titulo)
        c.fill = _fill(C_CINZA_MED); c.font = _font(bold=True, color=C_BRANCO, size=9)
        c.alignment = _align('center' if col > 1 else 'left', 'center')
        c.border = _border()
    row += 1

    # Linhas de dados
    linhas = [
        ('Total de Lançamentos Analisados',      total,       '100%',           None),
        ('Total de Inconsistências Encontradas', total_erros, f'{pct:.1f}%',    None),
    ]
    for sev_nome, cor_sev in [('CRÍTICO', C_CRITICO), ('ALTO', C_ALTO),
                               ('MÉDIO', C_MEDIO),   ('BAIXO', C_BAIXO)]:
        qtd_s = sev.get(sev_nome, 0)
        pct_s = (qtd_s / total_erros * 100) if total_erros > 0 else 0
        linhas.append((f'  └ {sev_nome}', qtd_s, f'{pct_s:.1f}%', (sev_nome, cor_sev)))

    for i, (ind, qtd, pct_str, sev_info) in enumerate(linhas):
        bg = C_CINZA_CLARO if i % 2 == 0 else C_BRANCO
        if sev_info:
            bg = SEV_BG.get(sev_info[0], bg)
        ws.row_dimensions[row].height = 16

        c1 = ws.cell(row=row, column=1, value=ind)
        c1.fill = _fill(bg); c1.alignment = _align('left', 'center'); c1.border = _border()
        c1.font = _font(color=sev_info[1] if sev_info else C_TEXTO, size=9)

        c2 = ws.cell(row=row, column=2, value=qtd)
        c2.fill = _fill(bg); c2.alignment = _align('center', 'center'); c2.border = _border()
        c2.font = _font(bold=bool(sev_info), color=sev_info[1] if sev_info else C_TEXTO, size=9)

        c3 = ws.cell(row=row, column=3, value=pct_str)
        c3.fill = _fill(bg); c3.font = _font(size=9)
        c3.alignment = _align('center', 'center'); c3.border = _border()
        row += 1

    row += 1  # spacer

    # ── Seção: Por Regra ──────────────────────────────────────────────────────
    if por_regra:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        ws.row_dimensions[row].height = 24
        c = ws.cell(row=row, column=1, value='  INCONSISTÊNCIAS POR TIPO DE REGRA')
        c.fill = _fill(C_CINZA_ESC); c.font = _font(bold=True, color=C_BRANCO, size=10)
        c.alignment = _align('left', 'center')
        row += 1

        ws.row_dimensions[row].height = 18
        for col, titulo in [(1, 'Regra de Auditoria'), (2, 'Ocorrências'), (3, '% das Inconsist.')]:
            c = ws.cell(row=row, column=col, value=titulo)
            c.fill = _fill(C_CINZA_MED); c.font = _font(bold=True, color=C_BRANCO, size=9)
            c.alignment = _align('center' if col > 1 else 'left', 'center')
            c.border = _border()
        row += 1

        for i, (nome_r, qtd) in enumerate(sorted(por_regra.items(), key=lambda x: -x[1])):
            pct_r = (qtd / total_erros * 100) if total_erros > 0 else 0
            bg = C_CINZA_CLARO if i % 2 == 0 else C_BRANCO
            ws.row_dimensions[row].height = 16

            c1 = ws.cell(row=row, column=1, value=nome_r)
            c1.fill = _fill(bg); c1.font = _font(size=9)
            c1.alignment = _align('left', 'center'); c1.border = _border()

            c2 = ws.cell(row=row, column=2, value=qtd)
            c2.fill = _fill(bg); c2.font = _font(size=9)
            c2.alignment = _align('center', 'center'); c2.border = _border()

            c3 = ws.cell(row=row, column=3, value=f'{pct_r:.1f}%')
            c3.fill = _fill(bg); c3.font = _font(size=9)
            c3.alignment = _align('center', 'center'); c3.border = _border()
            row += 1


# ─── Aba por regra ────────────────────────────────────────────────────────────

def _criar_aba_regra(wb, nome_regra, erros, source_path='', source_sheet_name=''):
    ws = wb.create_sheet(_tab_name(nome_regra))
    ws.sheet_view.showGridLines = False

    # Larguras: Sev | Fornecedor | CNPJ | Linha | Aba origem | Link origem | Valor | Emissão | Vencimento | Categoria | Evidência
    for letra, w in zip('ABCDEFGHIJK', [14, 34, 22, 10, 22, 16, 16, 16, 16, 24, 38]):
        ws.column_dimensions[letra].width = w

    row = 1

    # Cabeçalho da aba
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
    ws.row_dimensions[row].height = 32
    qtd = len(erros)
    c = ws.cell(row=row, column=1,
                value=f'{nome_regra.upper()}  —  {qtd} ocorrência{"s" if qtd != 1 else ""}')
    c.fill = _fill(C_CINZA_ESC); c.font = _font(bold=True, color=C_VERDE, size=12)
    c.alignment = _align('left', 'center')
    row += 1

    # Caixa de explicação (3 colunas)
    expl = EXPLICACOES.get(nome_regra, {})
    if expl:
        # Títulos
        ws.row_dimensions[row].height = 16
        for col_s, col_e, titulo in [(1, 4, 'O QUE É'), (5, 7, 'POR QUE IMPORTA'), (8, 11, 'COMO CORRIGIR')]:
            ws.merge_cells(start_row=row, start_column=col_s, end_row=row, end_column=col_e)
            c = ws.cell(row=row, column=col_s, value=titulo)
            c.fill = _fill(C_CINZA_CLARO); c.font = _font(bold=True, color=C_CINZA_ESC, size=8)
            c.alignment = _align('left', 'center'); c.border = _border()
        row += 1

        # Conteúdo
        ws.row_dimensions[row].height = 72
        for col_s, col_e, chave in [(1, 4, 'o_que_e'), (5, 7, 'por_que_importa'), (8, 11, 'como_corrigir')]:
            ws.merge_cells(start_row=row, start_column=col_s, end_row=row, end_column=col_e)
            c = ws.cell(row=row, column=col_s, value=expl.get(chave, ''))
            c.fill = _fill('FAFAFA'); c.font = _font(size=8, color=C_TEXTO)
            c.alignment = _align('left', 'top', wrap=True); c.border = _border()
        row += 2  # spacer

    # Cabeçalho da tabela
    cabecalhos = ['Severidade', 'Fornecedor / Cliente', 'CNPJ/CPF',
                  'Linha (arquivo)', 'Aba origem', 'Link origem', 'Valor',
                  'Data Emissão (arquivo)', 'Data Vencimento (arquivo)', 'Categoria', 'Evidência']
    ws.row_dimensions[row].height = 18
    for col_idx, cab in enumerate(cabecalhos, 1):
        c = ws.cell(row=row, column=col_idx, value=cab)
        c.fill = _fill(C_CINZA_ESC); c.font = _font(bold=True, color=C_BRANCO, size=9)
        c.alignment = _align('center' if col_idx != 2 else 'left', 'center')
        c.border = _border()
    ws.freeze_panes = ws.cell(row=row + 1, column=1)
    row += 1

    # Dados
    for erro in erros:
        sev_nome = erro.get('severidade', '')
        bg       = SEV_BG.get(sev_nome, C_BRANCO)
        cor_sev  = SEV_COR.get(sev_nome, C_TEXTO)

        fornecedor = str(erro.get('fornecedor', 'N/D'))
        cnpj       = str(erro.get('cnpj', 'N/D'))
        linha      = erro.get('linha_arquivo', 'N/D')
        aba_origem = str(erro.get('aba_arquivo_origem', 'N/D'))
        data_e     = str(erro.get('data_emissao', 'N/D'))
        data_v     = str(erro.get('data_vencimento', 'N/D'))
        categoria  = str(erro.get('categoria', 'N/D'))
        detalhe    = str(erro.get('detalhe', ''))

        try:
            valor = float(erro.get('valor', 0))
        except (TypeError, ValueError):
            valor = str(erro.get('valor', ''))

        ws.row_dimensions[row].height = 28

        defs = [
            (1, sev_nome,  'center', _font(bold=True, color=cor_sev, size=9), False),
            (2, fornecedor,'left',   _font(size=9), True),
            (3, cnpj,      'center', _font(size=9), False),
            (4, linha,     'center', _font(size=9), False),
            (5, aba_origem,'left',   _font(size=9), False),
            (6, '',        'center', _font(size=9), False),
            (7, valor,     'right',  _font(size=9), False),
            (8, data_e,    'center', _font(size=9), False),
            (9, data_v,    'center', _font(size=9), False),
            (10, categoria,'left',   _font(size=9), True),
            (11, detalhe,  'left',   _font(size=9), True),
        ]
        for col_idx, val, h_align, fnt, wrap in defs:
            c = ws.cell(row=row, column=col_idx, value=val)
            c.fill = _fill(bg); c.font = fnt; c.border = _border()
            c.alignment = _align(h_align, 'center' if not wrap else 'top', wrap=wrap)

        # Link direto para a linha no arquivo de origem (quando disponível)
        linha_int = linha if isinstance(linha, int) else None
        if linha_int is None:
            try:
                linha_int = int(linha)
            except (TypeError, ValueError):
                linha_int = None
        if source_sheet_name and linha_int:
            addr = f"#'{source_sheet_name}'!A{linha_int}"
            cell_link = ws.cell(row=row, column=6, value='Abrir linha')
            cell_link.hyperlink = addr
            cell_link.style = 'Hyperlink'
            cell_link.fill = _fill(bg)
            cell_link.border = _border()
            cell_link.alignment = _align('center', 'center')

        # Formato moeda para a coluna Valor
        if isinstance(valor, float):
            ws.cell(row=row, column=7).number_format = '#,##0.00'

        row += 1


# ─── Função principal ─────────────────────────────────────────────────────────

def gerar_excel(resultado, output_path, nome_arquivo_origem='', empresa='', source_path=''):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove aba vazia padrão

    source_sheet_name = _criar_aba_base_analisada(wb, source_path)
    _criar_resumo(wb, resultado, nome_arquivo_origem, empresa)

    erros = resultado.get('erros', [])
    erros_por_regra = defaultdict(list)
    for e in erros:
        erros_por_regra[e['regra']].append(e)

    for nome_regra in sorted(erros_por_regra, key=lambda r: -len(erros_por_regra[r])):
        _criar_aba_regra(
            wb,
            nome_regra,
            erros_por_regra[nome_regra],
            source_path=source_path,
            source_sheet_name=source_sheet_name or '',
        )

    wb.save(output_path)
    return output_path
